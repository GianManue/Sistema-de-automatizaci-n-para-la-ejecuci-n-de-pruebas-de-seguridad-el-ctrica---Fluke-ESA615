import csv
import openpyxl
from openpyxl.styles import Font
import os
import sys

#Este código recopila todos los csv y manda dicha información al excel del informe
class ProcesadorTotalExcel:
    def __init__(self, rutas_csv_informe, rutas_csv_datos, rutas_plantilla, ruta_directorio_salida):
        self.rutas_csv_informe = rutas_csv_informe
        self.rutas_csv_datos = rutas_csv_datos
        self.rutas_plantilla = rutas_plantilla
        self.ruta_directorio_salida = ruta_directorio_salida

        self.datos_informe = {}
        self.datos_bloques = {}

    def buscar_archivo(self, lista_rutas):
        """Busca un archivo y devuelve su ruta real."""
        for ruta in lista_rutas:
            ruta_norm = os.path.normpath(ruta)
            if os.path.exists(ruta_norm):
                return ruta_norm
        return None

    def cargar_csv_informe(self):
        ruta = self.buscar_archivo(self.rutas_csv_informe)
        if not ruta:
            print(f"\n[ERROR] No se encontró 'resultados_esa615.csv'.")
            print("Se buscaron en las siguientes rutas:")
            for r in self.rutas_csv_informe:
                print(f" - {os.path.normpath(r)}")
            return False

        try:
            with open(ruta, mode='r', encoding='utf-8') as archivo:
                lector = csv.DictReader(archivo)
                for fila_cruda in lector:
                    fila = dict(fila_cruda)
                    codigo = fila.get('Codigo', '').strip()
                    if codigo:
                        try:
                            self.datos_informe[codigo] = float(fila['Promedio'])
                        except:
                            pass
            print(f"[OK] {len(self.datos_informe)} datos cargados para INFORME desde:\n -> {ruta}")
            return True
        except Exception as e:
            print(f"Error al leer CSV Informe: {e}")
            return False

    def cargar_csv_datos(self):
        ruta = self.buscar_archivo(self.rutas_csv_datos)
        if not ruta:
            print(f"\n[ERROR] No se encontró 'resultados_bloques_esa615.csv'.")
            print("Se buscaron en las siguientes rutas:")
            for r in self.rutas_csv_datos:
                print(f" - {os.path.normpath(r)}")
            return False

        try:
            with open(ruta, mode='r', encoding='utf-8') as archivo:
                lector = csv.DictReader(archivo)
                for fila_cruda in lector:
                    fila = dict(fila_cruda)
                    codigo = fila.get('Codigo', '').strip()
                    if codigo:
                        try:
                            self.datos_bloques[codigo] = float(fila['Valor'])
                        except:
                            pass
            print(f"[OK] {len(self.datos_bloques)} datos cargados para DATOS desde:\n -> {ruta}")
            return True
        except Exception as e:
            print(f"Error al leer CSV Datos: {e}")
            return False

    def inyectar_en_hoja(self, wb, nombre_hoja, diccionario_datos):
        if nombre_hoja not in wb.sheetnames:
            print(f"[ERROR] La pestaña '{nombre_hoja}' no existe.")
            return 0

        hoja = wb[nombre_hoja]
        reemplazos = 0

        for fila in hoja.iter_rows():
            for celda in fila:
                if isinstance(celda.value, str):
                    codigo_celda = celda.value.strip()
                    if codigo_celda in diccionario_datos:
                        celda.value = diccionario_datos[codigo_celda]

                        if celda.font:
                            fuente = celda.font
                            celda.font = Font(color="000000", name=fuente.name, size=fuente.size, bold=fuente.bold,
                                              italic=fuente.italic)
                        else:
                            celda.font = Font(color="000000")

                        reemplazos += 1
        return reemplazos

    def generar_excel_final(self):
        ruta_maestra = self.buscar_archivo(self.rutas_plantilla)

        if not ruta_maestra:
            print(f"\n[ERROR] No se encontró la plantilla de Excel.")
            raise Exception("Plantilla no encontrada")

        print(f"\nAbriendo plantilla maestra desde:\n -> {ruta_maestra}")
        try:
            import warnings
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

            wb = openpyxl.load_workbook(ruta_maestra, data_only=False)

            if self.datos_informe:
                r1 = self.inyectar_en_hoja(wb, "Informe", self.datos_informe)
                print(f"  -> {r1} reemplazos en hoja 'Informe'.")

            if self.datos_bloques:
                r2 = self.inyectar_en_hoja(wb, "Datos", self.datos_bloques)
                print(f"  -> {r2} reemplazos en hoja 'Datos'.")

            os.makedirs(self.ruta_directorio_salida, exist_ok=True)
            ruta_salida = os.path.join(self.ruta_directorio_salida, "INFORME_COMPLETO_ESA615.xlsx")

            wb.save(ruta_salida)
            print(f"\n ¡DOCUMENTO ACTUALIZADO CON ÉXITO!")
            print(f"El informe está listo en:\n -> {ruta_salida}")

        except Exception as e:
            print(f"\n [ERROR FATAL AL GUARDAR] {e}")
            raise e


def ejecutar():
    print(f"\n--- MÓDULO DATA_TO_EXCEL (Búsqueda Exhaustiva) ---")

    directorios_base = []

    if getattr(sys, 'frozen', False):
        #Escaneamos absolutamente todas las posibles carpetas donde Windows corre el proceso
        directorios_base.append(os.path.dirname(sys.executable))  # 1. Donde está el .exe físicamente
        directorios_base.append(os.getcwd())  # 2. El directorio de trabajo actual
        directorios_base.append(os.path.dirname(sys.argv[0]))  # 3. La ruta de origen de la llamada
        directorio_recursos_internos = sys._MEIPASS  # 4. Donde están los recursos ocultos
    else:
        base_dev = os.path.dirname(os.path.abspath(__file__))
        directorios_base.append(base_dev)
        directorios_base.append(os.path.dirname(base_dev))
        directorio_recursos_internos = base_dev

    #Limpiamos rutas duplicadas
    directorios_base = list(dict.fromkeys([os.path.abspath(d) for d in directorios_base if d]))

    rutas_csv_informe = []
    rutas_csv_datos = []

    #Armar la lista de todos los posibles lugares donde podrían estar los CSV
    for base in directorios_base:
        carpeta_gen = os.path.join(base, "ArchivosGenerados")
        rutas_csv_informe.append(os.path.join(carpeta_gen, "resultados_esa615.csv"))
        rutas_csv_datos.append(os.path.join(carpeta_gen, "resultados_bloques_esa615.csv"))

    #Armar lista de dónde podría estar la plantilla Excel
    rutas_plantilla = [os.path.join(directorio_recursos_internos, "Copia de informe_pruebas.xlsx")]
    for base in directorios_base:
        rutas_plantilla.append(os.path.join(base, "Copia de informe_pruebas.xlsx"))

    #Asumimos inicialmente la primera ruta
    ruta_directorio_salida = os.path.join(directorios_base[0], "ArchivosGenerados")

    procesador = ProcesadorTotalExcel(rutas_csv_informe, rutas_csv_datos, rutas_plantilla, ruta_directorio_salida)

    cargo_inf = procesador.cargar_csv_informe()
    cargo_dat = procesador.cargar_csv_datos()

    if cargo_inf or cargo_dat:
        #Forzamos a que el Excel se guarde exactamente en la misma ruta donde sí existían los CSV
        ruta_csv_real = procesador.buscar_archivo(rutas_csv_informe) or procesador.buscar_archivo(rutas_csv_datos)
        if ruta_csv_real:
            procesador.ruta_directorio_salida = os.path.dirname(ruta_csv_real)

        procesador.generar_excel_final()
    else:
        print("\n[!] ERROR CRÍTICO: No se encontraron los CSV.")
        raise Exception("Archivos CSV no encontrados.")


if __name__ == "__main__":
    ejecutar()